#user.py

from openpyxl import load_workbook, Workbook

c_name = 1
c_id = 2
c_money = 3
c_lvl = 4

default_money = 10000

wb = load_workbook("userDB.xlsx")
ws = wb.active

def signup(_name, _id):
    ws.cell(row=2, column=c_name, value=_name)
    ws.cell(row=2, column=c_id, value =_id)
    ws.cell(row=2, column=c_money, value = default_money)
    ws.cell(row=2, column=c_lvl, value = 1)

    wb.save("userDB.xlsx")
    def checkRow():
    for row in range(2, ws.max_row + 1):
        if ws.cell(row,1).value is None:
            return row
            break
    #return ws.max_row 불안정하지만 훨씬 간단함

def signup(_name, _id):
    _row = checkRow()
    
    ws.cell(row=_row, column=c_name, value=_name)
    ws.cell(row=_row, column=c_id, value =_id)
    ws.cell(row=_row, column=c_money, value = default_money)
    ws.cell(row=_row, column=c_lvl, value = 1)
    
    def checkName(_name, _id):
    for row in range(2, ws.max_row+1):
        if ws.cell(row,1).value == _name and ws.cell(row,2).value == _id:
            break
            return False
        else:
            return True
            break
def delete():
ws.delete_rows(2,ws.max_row)
wb.save("userDB.xlsx")
def userInfo(_namd, _id):
	if not checkName(_namd, _id):
    	for row in range(2, ws.max_row+2):
    		if ws.cell(row, 1).value == _name and ws.cell(row, 2).value == _id:
        		return ws.cell(row,1).value, ws.cell(row,c_lvl).value
                break
    else:
    	return None, None
        def loadFile():
	wb.load_workbook("userDB.xlsx")
    ws= wb.active
    def _func():
    loadFile()
    ...
    wb.save("userDB.xlsx")
    wb.close()
    def getMoney(_name, _id):
	loadFile()
    
    for row in range(2, ws.max_row+2):
    	if ws.cell(row, c_name).value == _name and ws.cell(row,c_id).value == hex(_id):
        	return ws.cell(row,c_money).value
            break
        else:
        	return 0
            break
        def remit(sender, s_id, receiver, r_id, _amount):
    loadFile()
    
    receiver_row = findRow(receiver, r_id)
    sender_row = findRow(sender, s_id)
    
    ws.cell(receiver_row, c_money).value += int(_amount)
    ws.cell(sender_row, c_money).value -= int(_amount)

    saveFile()
c_loss = 5
def addLoss(_target, _row, _amount):
    loadFile()
    
    ws.cell(_row, c_loss).value += _amount

    saveFile()
    def Signup(_name, _id):
    loadFile()

    _row = checkFirstRow()

    ws.cell(row=_row, column=c_name, value=_name)
    ws.cell(row=_row, column=c_id, value =hex(_id))
    ws.cell(row=_row, column=c_money, value = default_money)
    ws.cell(row=_row, column=c_lvl, value = 1)
    ws.cell(row=_row, column=c_loss, value = 0)

    saveFile()
    def userInfo(_row):
    loadFile()

    _lvl = ws.cell(_row,c_lvl).value
    _money = ws.cell(_row,c_money).value
    _loss = ws.cell(_row,c_loss).value

    saveFile()

    return _lvl, _money, _loss
    c_name = 1
c_id = 2
c_lvl = 3
c_exp = 4
c_money = 5
c_loss = 6
def Signup(_name, _id):
    loadFile()

    _row = checkFirstRow()

    ws.cell(row=_row, column=c_name, value=_name)
    ws.cell(row=_row, column=c_id, value =hex(_id))

    ws.cell(row=_row, column=c_lvl, value = 1)
    ws.cell(row=_row, column=c_exp, value = 0)

    ws.cell(row=_row, column=c_money, value = default_money)
    ws.cell(row=_row, column=c_loss, value = 0)

    saveFile()
    def userInfo(_row):
    loadFile()

    _lvl = ws.cell(_row,c_lvl).value
    _exp = wsl.cell(_row,c_exp).value
    _money = ws.cell(_row,c_money).value
    _loss = ws.cell(_row,c_loss).value

    saveFile()

    return _lvl, _exp, _money, _loss
    def addExp(_row, _amount):
    loadFile()

    ws.cell(_row, c_exp).value += _amount

    saveFile()
def levelupCheck(_row):
    print("user.py - levelupCheck")
    loadFile()

    name = ws.cell(_row, c_name).value
    exp = ws.cell(_row, c_exp).value
    lvl = ws.cell(_row, c_lvl).value
    amount_to_up = lvl*lvl + 6*lvl

    if exp >= amount_to_up:
        ws.cell(_row, c_lvl).value += 1
        ws.cell(_row, c_exp).value -= amount_to_up
        return True, lvl+1
    else:
        return False, lvl
def addExp(_row, _amount):
    loadFile()

    ws.cell(_row, c_exp).value += _amount

    saveFile()
def levelupCheck(_row):
    loadFile()

    name = ws.cell(_row, c_name).value
    exp = ws.cell(_row, c_exp).value
    lvl = ws.cell(_row, c_lvl).value
    amount_to_up = lvl*lvl + 6*lvl
    count = 0

    if exp >= amount_to_up:
        while(exp >= amount_to_up and exp >= 0):
            ws.cell(_row, c_lvl).value += 1
            count += 1
            ws.cell(_row, c_exp).value -= amount_to_up

            lvl = ws.cell(_row, c_lvl).value
            exp = ws.cell(_row, c_exp).value
            amount_to_up = lvl*lvl + 6*lvl
        return True, lvl+count
    else:
        return False, lvl
def ranking():
    loadFile()

    userRanking =  {}
    userNum = checkUserNum()

    for row in range(2, 2+userNum):
        _name = ws.cell(row, c_name).value
        _lvl = ws.cell(row, c_lvl).value
        userRanking[_name] = _lvl
        
    a = sorted(userRanking.items(), reverse=True, key=lambda item:item[1])
    result = []
    for items in a:
        result.append(items[0])
        result.append(items[1])
    return result

def getRank(_row):
    print("user.py - getRank")
    user = ws.cell(_row, c_name).value
    rank = ranking()

    result = int(rank.index(user)/2)+1

    return result